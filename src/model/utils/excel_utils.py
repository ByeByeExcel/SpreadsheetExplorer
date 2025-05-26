import re
from typing import Iterable

import openpyxl.utils
import xlwings as xw

from model.domain_model.spreadsheet.range_reference import RangeReference, RangeReferenceType


def convert_to_absolute_range(address: str) -> str:
    if ":" in address:
        start, end = address.split(":")
        return f"{convert_to_absolute_address(start)}:{convert_to_absolute_address(end)}"
    return f"{convert_to_absolute_address(address)}"


def convert_to_absolute_address(address: str) -> str:
    return openpyxl.utils.absolute_coordinate(address)


def replace_cell_reference_in_formula(formula: str, target_cell: str, new_name: str) -> str:
    min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(target_cell)
    min_col_str = openpyxl.utils.cell.get_column_letter(min_col)
    max_col_str = openpyxl.utils.cell.get_column_letter(max_col)

    if min_col == max_col and min_row == max_row:
        pattern = rf"(?<![A-Z0-9$'])" \
                  rf"(\$?{min_col_str}\$?{min_row})" \
                  rf"(?![A-Z0-9])"
    else:
        pattern = rf"(?<![A-Z0-9$'])" \
                  rf"((?:'[^']+'|[A-Za-z0-9_]+)!?)?" \
                  rf"\$?{min_col_str}\$?{min_row}:\$?{max_col_str}\$?{max_row}" \
                  rf"(?![A-Z0-9])"

    return re.sub(pattern, new_name, formula, flags=re.IGNORECASE)


def convert_xlwings_address(cell_range: xw.Range) -> RangeReference:
    range_type: RangeReferenceType = RangeReferenceType.CELL if cell_range.shape == (1, 1) else RangeReferenceType.RANGE

    return RangeReference.from_raw(cell_range.sheet.book.name, cell_range.sheet.name, cell_range.address, range_type)


def ref_string_is_range(range_reference: str) -> bool:
    (min_col, min_row, max_col, max_row) = openpyxl.utils.range_boundaries(range_reference)
    return min_col != max_col or min_row != max_row


def get_address_from_offset(top_left_row, top_left_col, row, col) -> str:
    return get_address(top_left_row + row, top_left_col + col)


def get_address(row, col) -> str:
    return f"{openpyxl.utils.get_column_letter(col)}{row}"


def get_cell_references_of_range(range_ref: RangeReference) -> Iterable[RangeReference]:
    if range_ref.reference_type != RangeReferenceType.RANGE:
        raise ValueError("Range reference must be of type RANGE.")

    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(range_ref.reference)

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            yield RangeReference.from_raw(range_ref.workbook, range_ref.sheet, get_address(row, col))


def parse_range_reference(raw_ref: str, current_workbook: str, current_sheet: str) -> RangeReference:
    match = re.match(
        r"""^(?:\[(?P<workbook>[^\]]+)\])?  # Optional [Workbook]
             (?:(?P<sheet>'[^']+'|[^!]+)!)?         # Optional Sheet!
             (?P<address>                           # Required part
                 \$?[A-Z]{1,3}\$?\d+                # A1
                 (:\$?[A-Z]{1,3}\$?\d+)?            # or :B2
                 |
                 [A-Za-z_][A-Za-z0-9_]*             # or defined name
             )$""",
        raw_ref.strip(),
        re.VERBOSE | re.IGNORECASE
    )

    if not match:
        raise ValueError(f"Invalid Excel reference: '{raw_ref}'")

    workbook = match.group("workbook") or current_workbook
    sheet = match.group("sheet")
    address = match.group("address")

    if sheet:
        sheet = sheet.strip("'")
    else:
        sheet = current_sheet

    # Detect reference type
    if workbook.lower() != current_workbook.lower():
        ref_type = RangeReferenceType.EXTERNAL
    elif ":" in address:
        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(address)
        if min_col == max_col and min_row == max_row:
            address = get_address(min_row, min_col)
            ref_type = RangeReferenceType.CELL
        else:
            ref_type = RangeReferenceType.RANGE
    elif re.match(r"^\$?[A-Z]{1,3}\$?\d+$", address):
        ref_type = RangeReferenceType.CELL
    else:
        ref_type = RangeReferenceType.DEFINED_NAME

    return RangeReference.from_raw(workbook, sheet, address, ref_type)
